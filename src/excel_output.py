"""Excel output generation with MLF tables, heatmaps, and biggest movers."""

import logging
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from . import config

logger = logging.getLogger(__name__)

# Styling constants
HEADER_FONT = Font(bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def generate_all_workbooks(summary: pd.DataFrame, output_dir: str):
    """Generate one .xlsx workbook per region from the summary DataFrame."""
    output_path = Path(output_dir)
    output_path.mkdir(parents=True, exist_ok=True)

    # Export FY columns only (exclude import companions)
    fy_cols = sorted([
        c for c in summary.columns
        if c.startswith("FY") and "Import" not in c
    ])

    # Only include import FY columns with meaningful coverage (5+ DUIDs)
    # — avoids cluttering output with single-asset pumped-hydro import columns from early FYs
    import_fy_cols = sorted([
        c for c in summary.columns
        if c.endswith(" Import") and c.startswith("FY")
        and summary[c].notna().sum() >= 5
    ])

    for region in config.REGIONS:
        region_data = summary[summary["REGIONID"] == region].copy()
        if region_data.empty:
            logger.warning(f"No data for {region}, skipping workbook")
            continue

        friendly_name = config.REGION_NAMES[region]
        filepath = output_path / f"{friendly_name}_mlf.xlsx"
        _write_region_workbook(region_data, friendly_name, fy_cols, filepath,
                               import_fy_cols=import_fy_cols)
        logger.info(f"Written {filepath}")


def _write_region_workbook(data: pd.DataFrame, region_name: str,
                           fy_cols: list[str], filepath: Path,
                           import_fy_cols: list[str] | None = None):
    """Write a 3-sheet workbook for a single region."""
    wb = Workbook()

    _write_mlf_table(wb, data, region_name, fy_cols, import_fy_cols=import_fy_cols or [])
    _write_heatmap(wb, data, region_name, fy_cols, import_fy_cols=import_fy_cols or [])
    _write_movers(wb, data, region_name)

    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    wb.save(filepath)


IMPORT_HEADER_FILL = PatternFill(start_color="7B5EA7", end_color="7B5EA7", fill_type="solid")


def _build_fy_column_order(fy_cols: list[str], import_fy_cols: list[str]) -> list[str]:
    """Interleave import columns after their corresponding export FY column.

    E.g. [..., FY25-26, FY25-26 Import, FY26-27, FY26-27 Import, FY27-28 (Draft), ...]
    """
    import_set = set(import_fy_cols)
    ordered = []
    for fy in fy_cols:
        ordered.append(fy)
        companion = f"{fy} Import"
        # Draft columns have the pattern "FYxx-xx (Draft) Import"
        if "(Draft)" in fy:
            companion = f"{fy} Import"
        if companion in import_set:
            ordered.append(companion)
    return ordered


def _write_mlf_table(wb: Workbook, data: pd.DataFrame, region_name: str,
                     fy_cols: list[str], import_fy_cols: list[str] | None = None):
    """Sheet 1: Clean MLF table — all DUIDs with FY columns."""
    ws = wb.create_sheet(title="MLF Table")

    all_fy_cols = _build_fy_column_order(fy_cols, import_fy_cols or [])

    # Build headers
    meta_headers = ["DUID", "Station", "Fuel Type", "Capacity (MW)"]
    headers = meta_headers + all_fy_cols
    if "YOY_CHANGE" in data.columns:
        headers += ["Export YoY", "Export YoY %"]
    if "IMPORT_YOY_CHANGE" in data.columns:
        headers += ["Import YoY", "Import YoY %"]

    # Write headers
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = IMPORT_HEADER_FILL if "Import" in header else HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER

    # Sort by latest MLF ascending (worst first)
    sort_col = fy_cols[-1] if fy_cols else None
    if sort_col and sort_col in data.columns:
        data = data.sort_values(sort_col, na_position="last")

    # Write data
    for row_idx, (_, row) in enumerate(data.iterrows(), 2):
        ws.cell(row=row_idx, column=1, value=row.get("DUID", "")).border = THIN_BORDER
        ws.cell(row=row_idx, column=2, value=row.get("STATION_NAME", "")).border = THIN_BORDER
        ws.cell(row=row_idx, column=3, value=row.get("FUEL_CATEGORY", "")).border = THIN_BORDER

        cap = row.get("CAPACITY_MW")
        cell = ws.cell(row=row_idx, column=4, value=cap if pd.notna(cap) else "")
        cell.border = THIN_BORDER
        if pd.notna(cap):
            cell.number_format = "0.0"

        for col_offset, fy in enumerate(all_fy_cols):
            val = row.get(fy)
            cell = ws.cell(row=row_idx, column=5 + col_offset,
                           value=val if pd.notna(val) else "")
            cell.number_format = "0.0000"
            cell.alignment = Alignment(horizontal="center")
            cell.border = THIN_BORDER

        base_col = 5 + len(all_fy_cols)
        if "YOY_CHANGE" in data.columns:
            yoy = row.get("YOY_CHANGE")
            cell = ws.cell(row=row_idx, column=base_col,
                           value=yoy if pd.notna(yoy) else "")
            cell.number_format = "0.0000"
            cell.alignment = Alignment(horizontal="center")
            cell.border = THIN_BORDER

            yoy_pct = row.get("YOY_PCT_CHANGE")
            cell = ws.cell(row=row_idx, column=base_col + 1,
                           value=yoy_pct if pd.notna(yoy_pct) else "")
            cell.number_format = "0.00"
            cell.alignment = Alignment(horizontal="center")
            cell.border = THIN_BORDER
            base_col += 2

        if "IMPORT_YOY_CHANGE" in data.columns:
            iyoy = row.get("IMPORT_YOY_CHANGE")
            cell = ws.cell(row=row_idx, column=base_col,
                           value=iyoy if pd.notna(iyoy) else "")
            cell.number_format = "0.0000"
            cell.alignment = Alignment(horizontal="center")
            cell.border = THIN_BORDER

            iyoy_pct = row.get("IMPORT_YOY_PCT_CHANGE")
            cell = ws.cell(row=row_idx, column=base_col + 1,
                           value=iyoy_pct if pd.notna(iyoy_pct) else "")
            cell.number_format = "0.00"
            cell.alignment = Alignment(horizontal="center")
            cell.border = THIN_BORDER

    # Column widths
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 14
    for i in range(5, 5 + len(all_fy_cols) + 5):
        ws.column_dimensions[get_column_letter(i)].width = 13

    ws.freeze_panes = "E2"


def _write_heatmap(wb: Workbook, data: pd.DataFrame, region_name: str,
                   fy_cols: list[str], import_fy_cols: list[str] | None = None):
    """Sheet 2: MLF values with conditional colour formatting."""
    ws = wb.create_sheet(title="Heatmap")

    all_fy_cols = _build_fy_column_order(fy_cols, import_fy_cols or [])

    headers = ["DUID"] + all_fy_cols
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = IMPORT_HEADER_FILL if "Import" in header else HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER

    sort_col = fy_cols[-1] if fy_cols else None
    if sort_col and sort_col in data.columns:
        data = data.sort_values(sort_col, na_position="last")

    num_rows = len(data)
    for row_idx, (_, row) in enumerate(data.iterrows(), 2):
        ws.cell(row=row_idx, column=1, value=row.get("DUID", "")).border = THIN_BORDER
        for col_offset, fy in enumerate(all_fy_cols):
            val = row.get(fy)
            cell = ws.cell(row=row_idx, column=2 + col_offset,
                           value=val if pd.notna(val) else "")
            cell.number_format = "0.0000"
            cell.alignment = Alignment(horizontal="center")
            cell.border = THIN_BORDER

    # Apply colour scale: red (low MLF = bad) → yellow → green (high MLF = good)
    if num_rows > 0:
        for col_idx in range(2, 2 + len(all_fy_cols)):
            col_letter = get_column_letter(col_idx)
            cell_range = f"{col_letter}2:{col_letter}{num_rows + 1}"
            ws.conditional_formatting.add(
                cell_range,
                ColorScaleRule(
                    start_type="min", start_color="F8696B",   # red (low MLF)
                    mid_type="percentile", mid_value=50, mid_color="FFEB84",
                    end_type="max", end_color="63BE7B",       # green (high MLF)
                ),
            )

    ws.column_dimensions["A"].width = 14
    for i in range(2, 2 + len(all_fy_cols)):
        ws.column_dimensions[get_column_letter(i)].width = 13

    ws.freeze_panes = "B2"


def _write_movers(wb: Workbook, data: pd.DataFrame, region_name: str):
    """Sheet 3: Biggest movers — top degrading and improving assets."""
    ws = wb.create_sheet(title="Biggest Movers")

    if "YOY_CHANGE" not in data.columns:
        ws.cell(row=1, column=1, value="No YoY data available")
        return

    valid = data.dropna(subset=["YOY_CHANGE"]).copy()
    if valid.empty:
        ws.cell(row=1, column=1, value="No YoY data available")
        return

    # Top 20 degrading (most negative YoY change)
    degrading = valid.nsmallest(20, "YOY_CHANGE")
    # Top 20 improving (most positive YoY change)
    improving = valid.nlargest(20, "YOY_CHANGE")

    headers = ["DUID", "Station", "Fuel Type", "Latest MLF", "Prev MLF",
               "YoY Change", "YoY %"]

    def write_section(start_row: int, title: str, section_data: pd.DataFrame):
        ws.cell(row=start_row, column=1, value=title).font = Font(bold=True, size=13)
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=start_row + 1, column=col_idx, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center")
            cell.border = THIN_BORDER

        for row_idx, (_, row) in enumerate(section_data.iterrows(), start_row + 2):
            ws.cell(row=row_idx, column=1, value=row.get("DUID", "")).border = THIN_BORDER
            ws.cell(row=row_idx, column=2, value=row.get("STATION_NAME", "")).border = THIN_BORDER
            ws.cell(row=row_idx, column=3, value=row.get("FUEL_CATEGORY", "")).border = THIN_BORDER

            for ci, field, fmt in [(4, "LATEST_MLF", "0.0000"),
                                   (5, "PREV_MLF", "0.0000"),
                                   (6, "YOY_CHANGE", "0.0000"),
                                   (7, "YOY_PCT_CHANGE", "0.00")]:
                val = row.get(field)
                cell = ws.cell(row=row_idx, column=ci, value=val if pd.notna(val) else "")
                cell.number_format = fmt
                cell.alignment = Alignment(horizontal="center")
                cell.border = THIN_BORDER

        return start_row + 2 + len(section_data)

    next_row = write_section(1, f"{region_name} — Most Degraded (YoY)", degrading)
    write_section(next_row + 2, f"{region_name} — Most Improved (YoY)", improving)

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 12
    for i in range(4, 8):
        ws.column_dimensions[get_column_letter(i)].width = 13

    ws.freeze_panes = "A3"
